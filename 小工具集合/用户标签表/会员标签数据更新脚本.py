import pyodbc
import pyodbc
import pandas as pd
import datetime
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')
# 连接数据库
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'SERVER=192.168.0.169;'
                      'DATABASE=ChjBidb;'
                      'UID=chjreport;'
                      'PWD=Chj@12345;')
# SQL 查询
sql1 = """
    (select 
    '用户总量' AS 字段分类,
    '总用户数' as 字段名,
    count(distinct 会员ID) as 会员数量
    FROM BI_Business_Member)
    union all
    (select 
    '激活用户数' AS 字段分类,
    '激活总用户数' as 字段名,
    count(distinct 会员ID) as 会员数量
    FROM BI_Business_Member
    where [是否在云店上激活] = '是')
    union all
    (select 
    '婚恋标签' AS 字段分类,
    '婚恋标签总人数' as 字段名,
    count(distinct 会员号) as 会员数量
    FROM
    (SELECT distinct 会员号 as 会员号  FROM [dbo].[BI_Business_qywechatmember_tag] as q join BI_Business_Member as m on q.unionId = m.[微信unionId] where tageName like '%婚%' or tageName like '%恋%'
    union all
    SELECT distinct MemberCode as 会员号 FROM [dbo].[BI_Business_ExtendedAttr_all]  where AttrName = '婚恋状态'
    union ALL
    select 会员号 from BI_Business_Member where [婚姻状况] != '保密' OR [结婚纪念日] is not null) as t)
    union all
    (select 
    'RFM标签' AS 字段分类,
    'RFM标签总人数' as 字段名,
    count(distinct membercode) as 会员数量
    from
    BI_Business_ExtendedAttr
    where 
    AttrName = 'RFM')
    union all
    (select 
    '激活用户数' AS 字段分类,
    '激活总用户数' as 字段名,
    count(distinct 会员ID) as 会员数量
    FROM BI_Business_Member
    where [是否在云店上激活] = '是')
    union all
    (SELECT
    '年龄段' AS 字段分类,
    年代 AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 年代)
    UNION ALL
    (SELECT
    '性别' AS 字段分类,
    性别 AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 性别)
    UNION ALL
    (SELECT
    '等级' AS 字段分类,
    等级 AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 等级)
    UNION ALL
    (SELECT
    '婚姻状况' AS 字段分类,
    婚姻状况 AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 婚姻状况)
    UNION ALL
    (SELECT
    '来源' AS 字段分类,
    来源 AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 来源)
    UNION ALL
    (SELECT
    '会员绑定微信' AS 字段分类,
    CASE WHEN 是否绑定微信 = '1' THEN '已绑定' ELSE '未绑定' END AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 是否绑定微信)
    UNION ALL
    (SELECT
    '异常会员' AS 字段分类,
    CASE WHEN 是否异常会员 = '是' THEN '异常会员' ELSE '正常会员' END AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 是否异常会员)
    UNION ALL
    (SELECT
    '云店激活会员' AS 字段分类,
    CASE WHEN 是否在云店上激活 = '是' THEN '激活会员' ELSE '非激活会员' END AS 字段名,
    COUNT(*) AS 会员数量
    FROM BI_Business_Member
    GROUP BY 是否在云店上激活)
    UNION ALL
    (select 
    'RFM标签' AS 字段分类,
    AttrOptionName as 字段名,
    count(distinct membercode) as 会员数量
    from
    BI_Business_ExtendedAttr
    where 
    AttrName = 'RFM'
    GROUP BY AttrOptionName)
"""
sql2 = """
            WITH 消费统计 AS (
                SELECT 
                        cc.会员号,
                        COUNT(DISTINCT CASE WHEN cc.消费时间 BETWEEN DATEADD(yy, -1, '2023-05-10') AND '2023-05-10 23:59:59' THEN cc.单据编号 ELSE NULL END) AS num_purchase,
                        MAX(cc.消费时间) AS max_purchase
                    FROM BI_Business_Consume cc
                    WHERE cc.消费时间 <= '2023-05-10 23:59:59'
                    GROUP BY cc.会员号
                ),
                会员分类 AS (
                    SELECT
                        cv.会员号,
                        CASE WHEN a.num_purchase >= 1 AND cv.共消费次数 = 1 THEN '1' ELSE '0' END AS 首笔消费,
                        CASE WHEN a.num_purchase >= 1 AND cv.共消费次数 > 1 THEN '1' ELSE '0' END AS 活跃会员,
                        CASE WHEN a.max_purchase BETWEEN DATEADD(yy, -2, '2023-05-10') AND DATEADD(yy, -1, '2023-05-10 23:59:59') THEN '1' ELSE '0' END AS 瞌睡会员,
                        CASE WHEN a.max_purchase BETWEEN DATEADD(yy, -3, '2023-05-10') AND DATEADD(yy, -2, '2023-05-10 23:59:59') THEN '1' ELSE '0' END AS 沉睡会员,
                        CASE WHEN a.max_purchase <= DATEADD(yy, -3, '2023-05-10 23:59:59') THEN '1' ELSE '0' END AS 流失会员
                    FROM BI_Business_Member cv
                    LEFT JOIN 消费统计 a ON cv.会员号 = a.会员号
                )
                SELECT '活跃状态'as 字段分类,'首笔消费' AS 字段名, COUNT(distinct 会员分类.会员号) AS 会员数量 FROM 会员分类 WHERE 首笔消费 = '1'
                UNION ALL
                SELECT '活跃状态'as 字段分类,'活跃会员' AS 字段名, COUNT(distinct 会员分类.会员号) AS 会员数量 FROM 会员分类 WHERE 活跃会员 = '1'
                UNION ALL
                SELECT '活跃状态'as 字段分类,'瞌睡会员' AS 字段名, COUNT(distinct 会员分类.会员号) AS 会员数量 FROM 会员分类 WHERE 瞌睡会员 = '1'
                    UNION ALL
                    SELECT '活跃状态'as 字段分类,'沉睡会员' AS 字段名, COUNT(distinct 会员分类.会员号) AS 会员数量 FROM 会员分类 WHERE 沉睡会员 = '1'
                    UNION ALL
                    SELECT '活跃状态'as 字段分类,'流失会员' AS 字段名, COUNT(distinct 会员分类.会员号) AS 会员数量 FROM 会员分类 WHERE 流失会员 = '1'
        """
# 将查询结果读取到DataFrame中
df1 = pd.read_sql_query(sql1, conn)
df2 = pd.read_sql_query(sql2, conn)
# 将df1和df2合并
df = pd.concat([df1, df2])
# 关闭连接
conn.close()
df = pd.concat([df1, df2])
today = datetime.date.today()
# 将字段分类列改名为今天的日期
df.rename(columns={'会员数量': today}, inplace=True)
# 重新给索引排序
df = df.reset_index(drop=True)
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取用户画像表
wb = load_workbook(r'\\192.168.0.88\潮宏基集团\潮宏基珠宝\品牌管理中心\CRM管理部\客户服务\85 临时数据看板\用户标签表.xlsx')
# 读取用户画像表中的第一个sheet
ws = wb.worksheets[0]
# 获取用户画像表中的最大行数
max_row = ws.max_row
# 获取用户画像表中的最大列数
max_column = ws.max_column
print(max_row)
# 添加表头
ws.cell(row=1, column=max_column + 1, value=today)

# 将df中的tommorow列的数据增加到用户画像表中的最后一列之后
# 将df中的tommorow列的数据增加到用户画像表中的最后一列之后
for i, (_, row) in enumerate(df.iterrows(), start=1):
    ws.cell(row=i + 1, column=max_column + 1, value=row[today])
# 保存用户画像表
wb.save(r'\\192.168.0.88\潮宏基集团\潮宏基珠宝\品牌管理中心\CRM管理部\客户服务\85 临时数据看板\用户标签表.xlsx')

# 李欣真可爱
print("李欣真可爱")