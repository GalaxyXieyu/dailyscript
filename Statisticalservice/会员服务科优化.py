# -*- coding: utf-8 -*-
# 去除警告
import warnings
warnings.filterwarnings('ignore')
import pandas as pd
import pymssql 
import datetime
server = '192.168.0.169'
user = 'chjreport'
password = 'Chj@12345'
database = 'ChjBidb'
# 数据处理
def decode_columns(df, column_names):
    """
    对pandas数据框中的多个列进行编解码，并返回新的数据框
    
    Args:
        chuda: pandas数据框
        column_names: list，需要进行编解码的列名列表
    
    Returns:
        pandas数据框，编解码后的结果
    """
    for column_name in column_names:
        # 跳过空值
        mask = df[column_name].notnull()
        # 对指定列进行编解码
        df.loc[mask, column_name] = df.loc[mask, column_name].apply(lambda x: x.encode('latin1').decode('gbk'))
    # 返回编解码后的结果
    return df
# 动态时间
def get_dynamic_dates():
    # 获取今天的日期
    today = datetime.date.today()
    # 获取本月的第一天
    first_day_of_month = today.replace(day=1)
    # 判断今天是否在本月的3天内
    if today.day <4:
        # 如果是，将开始时间设置为上个月的1号
        start_date = (first_day_of_month - datetime.timedelta(days=1)).replace(day=1)
    else:
        # 如果不是，将开始时间设置为本月的1号
        start_date = first_day_of_month
    # 将结束时间设置为昨天
    end_date = today - datetime.timedelta(days=1)
    # 返回开始时间和结束时间
    return start_date, end_date
start_date,end_date=get_dynamic_dates()
# 日期参数生成
params = {"start_date":start_date,"end_date":end_date}
print(params)
# 四舍五入函数
round_func = lambda x: round(x, 2)
# 把重新赋值一下params中的start和enddate为5月1和5月3
with pymssql.connect(server, user, password,database) as conn:
    with conn.cursor(as_dict=True) as cursor:
        # 兑礼数据
        sql1 = f'''
                    select 
                        s.[大区名称],
                        s.[小区名称],
                        s.门店名称,
                        f.[新老会员],
                        count(distinct id) as 计划触达,
                        sum(case when f.revisitTime is null then 0 else 1 end) as 实际触达
                    FROM
                        (select t.*,
                            m.新老会员
                        from v_crm_shoptask_detail t 
                        left join v_crm_shoptask_detail_members m on t.taskNo = m.任务编号 and t.memberCode = m.会员号
                        where t.BrandCode='CHJ' and t.taskName like '%消费7天回访%' and m.新老会员 is not Null
                        and t.revisitGuideNo not in (select distinct guide_no from v_mall_guide_all where status='5')
                        and t.startTime between '{start_date}' and '{end_date}'+' 23:59:59') as f
                    JOIN
                        BI_Base_Shop as s
                    on 
                        s.[门店代码] = f.revisitShopCode
                    GROUP BY
                        s.[大区名称],
                        s.[小区名称],
                        s.门店名称,
                        f.[新老会员]
                '''
        sql2 = f'''
                    select  
                    --  t.门店代码,t.会员号,t.导购代码,t.触达任务名称,xs.BILLNO as 单据编号,xs.qty as 数量,xs.AMOUNT as 金额,t.新老会员
                    t.大区名称,t.小区名称,t.门店名称,t.新老会员,count(distinct t.会员号) as 转化人数, sum(xs.AMOUNT) as 转化金额
                    from v_orderdetail xs
                    left join v_btgoods sp on xs.GOODSCODE = sp.商品代码
                    join (
                    select 
                        s.[大区名称],
                        s.[小区名称],
                        s.门店名称,
                        t.revisitShopCode as 门店代码,
                        t.memberCode as 会员号,t.revisitGuideNo as 导购代码,
                        m.新老会员,
                        case when t.taskName like '%消费7天回访%' then '会员消费后第7天回访任务' when t.taskName like '%☆积分礼品到店领取通知%' then '☆积分礼品到店领取通知' else t.taskName end as 触达任务名称,
                        min(t.startTime) as 触达开始时间
                    from v_crm_shoptask_detail t
                    join BI_Base_Shop as s
                    on t.revisitShopCode = s.[门店代码]
                    left join v_crm_shoptask_detail_members m on t.taskNo = m.任务编号 and t.memberCode = m.会员号
                    where t.status='1'
                    and t.BrandCode='CHJ'
                    and t.startTime BETWEEN '{start_date}' AND '{end_date}'+' 23:59:59'
                    and t.revisitGuideNo not in (select distinct guide_no from v_mall_guide_all where status='5')
                    group by 
                        s.[大区名称],
                        s.[小区名称],
                        s.门店名称,
                        t.revisitShopCode,
                        t.memberCode,
                        t.revisitGuideNo,
                        case when t.taskName like '%消费7天回访%' then '会员消费后第7天回访任务' when t.taskName like '%☆积分礼品到店领取通知%' then '☆积分礼品到店领取通知' else t.taskName end,
                        m.新老会员
                    )t on xs.VIPCARDNO = t.会员号
                    where xs.BILLDATE between t.触达开始时间 and '{end_date}'+' 23:59:59'
                    and sp.二级大类 <>'促销物料' and t.触达任务名称 = '会员消费后第7天回访任务'
                    GROUP BY
                        t.大区名称,t.小区名称,t.门店名称,t.新老会员
                '''
        cursor.execute(sql1)
        chuda1 = pd.DataFrame(cursor.fetchall())
        cursor.execute(sql2)
        chuda2 = pd.DataFrame(cursor.fetchall())
        # 判断是否chuda2为空，若为空则生成一个空的df
        if chuda2.empty:
            chuda2 = pd.DataFrame(columns=['大区名称','小区名称','门店名称','新老会员','转化人数','转化金额'])
        chuda1 = decode_columns(chuda1,["大区名称","小区名称",'门店名称',"新老会员"])
        chuda2 = decode_columns(chuda2,["大区名称","小区名称",'门店名称',"新老会员"])
        chuda = chuda1.merge(chuda2,how='left',on = ['大区名称','小区名称','门店名称','新老会员'])
# 计算触达完成率
chuda['触达完成率'] = chuda['实际触达'] / chuda['计划触达']
# 计算转化率
chuda['转化率'] = chuda['转化人数'] / chuda['实际触达']
order = ['大区名称', '小区名称', '门店名称', '新老会员', '计划触达', '实际触达', '触达完成率','转化人数','转化率','转化金额']
chuda = chuda[order]
# 计算转化金额（万元）
# 按大区和小区分组计算指标
grouped = chuda.groupby(['大区名称', '小区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
# 计算触达完成率
grouped['触达完成率'] = grouped['实际触达'] / grouped['计划触达']
# 计算转化率
grouped['转化率'] = grouped['转化人数'] / grouped['实际触达']
# 计算转化金额（万元）
grouped['转化金额'] = grouped['转化金额'] 
# 按新老会员分类计算指标
new_members = chuda[chuda['新老会员'] == '新会员']
old_members = chuda[chuda['新老会员'] == '老会员']
# 按大区和小区分组计算新会员触达率、转化率和转化金额（万元）
new_grouped = new_members.groupby(['大区名称', '小区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
new_grouped['新会员触达率'] = new_grouped['实际触达'] / new_grouped['计划触达']
new_grouped['新会员转化率'] = new_grouped['转化人数'] / new_grouped['实际触达']
new_grouped['新会员转化金额'] = new_grouped['转化金额'] /10000
new_grouped = new_grouped.rename(columns={'计划触达': '新会员计划触达', '实际触达': '新会员实际触达','转化人数': '新会员转化人数','新会员转化金额': '新会员转化金额(万)'})
new_grouped = new_grouped[['新会员计划触达','新会员实际触达','新会员触达率','新会员转化人数','新会员转化率', '新会员转化金额(万)']]
# 按大区和小区分组计算老会员触达率、转化率和转化金额（万元）
old_grouped = old_members.groupby(['大区名称', '小区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
# 修改列名
old_grouped['老会员触达率'] = old_grouped['实际触达'] / old_grouped['计划触达']
old_grouped['老会员转化率'] = old_grouped['转化人数'] / old_grouped['实际触达']
old_grouped['老会员转化金额'] = old_grouped['转化金额']/10000
old_grouped = old_grouped.rename(columns={'计划触达': '老会员计划触达', '实际触达': '老会员实际触达','转化人数': '老会员转化人数','老会员转化金额': '老会员转化金额(万)'})
old_grouped = old_grouped[['老会员计划触达','老会员实际触达','老会员触达率','老会员转化人数','老会员转化率', '老会员转化金额(万)']]
# 转化单位
grouped["转化金额(万)"] = grouped["转化金额"]/10000
grouped.fillna('0',inplace=True)
old_grouped.fillna('0',inplace=True)
new_grouped.fillna('0',inplace=True)
# 合并指标
result = pd.concat([grouped, new_grouped, old_grouped], axis=1).reset_index()
order = ['大区名称', '小区名称', '计划触达', '实际触达', '触达完成率', '转化人数', '转化率', '转化金额(万)','新会员计划触达','新会员实际触达','新会员触达率','新会员转化人数','新会员转化率', '新会员转化金额(万)','老会员计划触达','老会员实际触达','老会员触达率','老会员转化人数','老会员转化率', '老会员转化金额(万)']
result1 = result[order]
result1[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']]=result1[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']].astype(float)
result1[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']]=result1[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']].applymap(lambda x: '{:.2%}'.format(x) )
result1[['转化金额(万)','新会员转化金额(万)','老会员转化金额(万)']]= result1[['转化金额(万)','新会员转化金额(万)','老会员转化金额(万)']].applymap(round_func)
# 对于需要汇总的列，使用 sum() 函数进行汇总
columns_to_sum = ['计划触达', '实际触达', '转化人数', '转化金额(万)',
                  '新会员计划触达', '新会员实际触达', '新会员转化人数', '新会员转化金额(万)',
                  '老会员计划触达', '老会员实际触达', '老会员转化人数', '老会员转化金额(万)']
columns_to_convert = ['转化金额(万)',
                  '新会员计划触达', '新会员实际触达', '新会员转化人数', '新会员转化金额(万)',
                  '老会员计划触达', '老会员实际触达', '老会员转化人数', '老会员转化金额(万)']
result1[columns_to_convert].fillna(0,inplace=True)
result1[columns_to_convert] = result1[columns_to_convert].astype(float)
result1.replace("nan%",'-',inplace=True)
result1.replace("0",'-',inplace=True)
for col in columns_to_sum:
    try:
        result1.loc['汇总', col] = result1[col].sum()
    except:
        result1.loc['汇总', col] = 0
result1.sort_values(by=["大区名称","小区名称"],inplace=True)
result1.sort_values(by='触达完成率',ascending=True,inplace=True)
# 计算汇总行的触达完成率、转化率、新老会员触达率和转化率
result1.loc['汇总', '触达完成率'] = str(round(result1.loc['汇总', '实际触达'] / result1.loc['汇总', '计划触达']*100,2))+"%"
result1.loc['汇总', '转化率'] = str(round(result1.loc['汇总', '转化人数'] / result1.loc['汇总', '实际触达']*100,2))+"%"
result1.loc['汇总', '新会员触达率'] = str(round(result1.loc['汇总', '新会员实际触达'] / result1.loc['汇总', '新会员计划触达']*100,2))+"%"
result1.loc['汇总', '新会员转化率'] = str(round(result1.loc['汇总', '新会员转化人数'] / result1.loc['汇总', '新会员实际触达']*100,2))+"%"
result1.loc['汇总', '老会员触达率'] = str(round(result1.loc['汇总', '老会员实际触达'] / result1.loc['汇总', '老会员计划触达']*100,2))+"%"
result1.loc['汇总', '老会员转化率'] = str(round(result1.loc['汇总', '老会员转化人数'] / result1.loc['汇总', '老会员实际触达']*100,2))+"%"
order = ["大区名称","小区名称",'计划触达', '实际触达', '触达完成率', '转化人数', '转化率', '转化金额(万)','新会员触达率','新会员转化率','新会员转化金额(万)','老会员触达率','老会员转化率', '老会员转化金额(万)']
result1_final = result1[order]
result1_final.fillna(0,inplace=True)
# 计算转化金额（万元）
# 按大区和小区分组计算指标
grouped = chuda.groupby(['大区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
# 计算触达完成率
grouped['触达完成率'] = grouped['实际触达'] / grouped['计划触达']

# 计算转化率
grouped['转化率'] = grouped['转化人数'] / grouped['实际触达']

# 计算转化金额（万元）
grouped['转化金额'] = grouped['转化金额']/10000  

# 按新老会员分类计算指标
new_members = chuda[chuda['新老会员'] == '新会员']
old_members = chuda[chuda['新老会员'] == '老会员']

# 按大区和小区分组计算新会员触达率、转化率和转化金额（万元）
new_grouped = new_members.groupby(['大区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
new_grouped['新会员触达率'] = new_grouped['实际触达'] / new_grouped['计划触达']
new_grouped['新会员转化率'] = new_grouped['转化人数'] / new_grouped['实际触达']
new_grouped['新会员转化金额'] = new_grouped['转化金额'] /10000
new_grouped = new_grouped.rename(columns={'计划触达': '新会员计划触达', '实际触达': '新会员实际触达','转化人数': '新会员转化人数','新会员转化金额': '新会员转化金额(万)'})
# 按大区和小区分组计算老会员触达率、转化率和转化金额（万元）
old_grouped = old_members.groupby(['大区名称']).agg({
    '计划触达': 'sum',
    '实际触达': 'sum',
    '转化人数': 'sum',
    '转化金额': 'sum'
})
# 修改列名
old_grouped['老会员触达率'] = old_grouped['实际触达'] / old_grouped['计划触达']
old_grouped['老会员转化率'] = old_grouped['转化人数'] / old_grouped['实际触达']
old_grouped['老会员转化金额'] = old_grouped['转化金额'] /10000
old_grouped = old_grouped.rename(columns={'计划触达': '老会员计划触达', '实际触达': '老会员实际触达','转化人数': '老会员转化人数','老会员转化金额': '老会员转化金额(万)'})
grouped = grouped.rename(columns={'转化金额':'转化金额(万)'})
# 合并指标
result2 = pd.concat([grouped, new_grouped, old_grouped], axis=1).reset_index()
order = ['大区名称','计划触达', '实际触达', '触达完成率', '转化人数', '转化率', '转化金额(万)','新会员计划触达','新会员实际触达','新会员触达率','新会员转化人数','新会员转化率', '新会员转化金额(万)','老会员计划触达','老会员实际触达','老会员触达率','老会员转化人数','老会员转化率', '老会员转化金额(万)']
result2 = result2[order]
result2[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']]=result2[['触达完成率','转化率','新会员触达率','新会员转化率','老会员触达率','老会员转化率']].applymap(lambda x: '{:.2%}'.format(x))
result2[['转化金额(万)','新会员转化金额(万)','老会员转化金额(万)']]= result2[['转化金额(万)','新会员转化金额(万)','老会员转化金额(万)']].applymap(round_func)
result2.replace("nan%",'-',inplace=True)
result2.fillna("-",inplace=True)
result2.loc['汇总'] = None
# 对于需要汇总的列，使用 sum() 函数进行汇总
columns_to_sum = ['计划触达', '实际触达', '转化人数', '转化金额(万)',
                  '新会员计划触达', '新会员实际触达', '新会员转化人数', '新会员转化金额(万)',
                  '老会员计划触达', '老会员实际触达', '老会员转化人数', '老会员转化金额(万)']
columns_to_convert = ['转化金额(万)',
                  '新会员计划触达', '新会员实际触达', '新会员转化人数', '新会员转化金额(万)',
                  '老会员计划触达', '老会员实际触达', '老会员转化人数', '老会员转化金额(万)']
import numpy as np
result2[columns_to_convert] = result2[columns_to_convert].replace('-', np.nan).astype(float)
for col in columns_to_sum:
    try:
        result2.loc['汇总', col] = result2[col].sum()
    except:
        result2.loc['汇总', col] = 0
result2.sort_values(by=["大区名称"],inplace=True)
result2.sort_values(by='触达完成率',ascending=True,inplace=True)
# 计算汇总行的触达完成率、转化率、新老会员触达率和转化率
result2.loc['汇总', '触达完成率'] = str(round(result2.loc['汇总', '实际触达'] / result2.loc['汇总', '计划触达']*100,2))+"%"
result2.loc['汇总', '转化率'] = str(round(result2.loc['汇总', '转化人数'] / result2.loc['汇总', '实际触达']*100,2))+"%"
result2.loc['汇总', '新会员触达率'] = str(round(result2.loc['汇总', '新会员实际触达'] / result2.loc['汇总', '新会员计划触达']*100,2))+"%"
result2.loc['汇总', '新会员转化率'] = str(round(result2.loc['汇总', '新会员转化人数'] / result2.loc['汇总', '新会员实际触达']*100,2))+"%"
result2.loc['汇总', '老会员触达率'] = str(round(result2.loc['汇总', '老会员实际触达'] / result2.loc['汇总', '老会员计划触达']*100,2))+"%"
result2.loc['汇总', '老会员转化率'] = str(round(result2.loc['汇总', '老会员转化人数'] / result2.loc['汇总', '老会员实际触达']*100,2))+"%"

result2.sort_values(by=["大区名称"],inplace=True)
order = ["大区名称",'计划触达', '实际触达', '触达完成率', '转化人数', '转化率', '转化金额(万)','新会员触达率','新会员转化率','新会员转化金额(万)','老会员触达率','老会员转化率', '老会员转化金额(万)']
result2_final = result2[order]

# 售后券转化情况

with pymssql.connect(server, user, password,database) as conn:
    with conn.cursor(as_dict=True) as cursor:
        # 兑礼数据
        sql2 = f'''
                    select 
                        t.no as 订单编号,t.date_add as 兑换时间,t.shopCode as 兑换门店,s.门店名称,s.大区名称,s.小区名称,
                        isnull(cv.会员号,t.tel) as 会员号,t.consignee as 会员名称,isnull(cv.等级,cv2.等级) as 等级,
                        t.sku as 券号,t.product_title as 券名,t.quantity as 数量,cp.billno as 服务编码,isnull(cp.billdate,r.is_use_time) as 门店受理时间,isnull(cp.优惠券号,r.优惠券号) as 优惠券号,cast(t.date_add as varchar(50)) as date_add_char
                    from v_mall_order_detail t
                    left join v_shop_dict s on t.shopCode = s.门店代码
                    left join BI_Business_Member cv on t.memberId = cv.会员ID
                    left join BI_Business_Member cv2 on t.tel = cv2.会员号
                    left join (
                        select a.子订单号,a.优惠券号,b.billno,b.billdate
                        from v_mall_sku_coupon_transf a
                        join BI_Coupons_servicebill b on a.优惠券号 = b.coupon
                    )cp on t.order_sku_no = cp.子订单号
                    left join (
                        select a.子订单号,a.优惠券号,c.is_use_time 
                        from v_mall_sku_coupon_transf a
                        join v_crm_coupon_detail c on a.优惠券号 = c.COUPON_CODE 
                    )r on t.order_sku_no = r.子订单号
                    where t.statu in (1,2,3,4)
                    and t.title='虚拟商品'
                    and t.showLocation like '%售后%'
                    and t.date_add>= '{start_date}'
                    AND t.date_add<= '{end_date}'+' 23:59:59'
                '''
        cursor.execute(sql2)
        shquan1 = decode_columns(pd.DataFrame(cursor.fetchall()),["等级"])
        sql3 = f'''
                select 
                    m.等级,
                    q.券名,
                    count(distinct q.会员号) as 会员数,
                    SUM(c.金额) as 转化金额
                FROM
                    (select 
                        isnull(cv.会员号,t.tel) as 会员号,t.product_title as 券名,min(t.date_add) 最先兑换时间
                    from v_mall_order_detail t
                    left join v_shop_dict s on t.shopCode = s.门店代码
                    left join BI_Business_Member cv on t.memberId = cv.会员ID
                    left join BI_Business_Member cv2 on t.tel = cv2.会员号
                    left join (
                        select a.子订单号,a.优惠券号,b.billno,b.billdate
                        from v_mall_sku_coupon_transf a
                        join BI_Coupons_servicebill b on a.优惠券号 = b.coupon
                    )cp on t.order_sku_no = cp.子订单号
                    where t.statu in (1,2,3,4)
                    and t.title='虚拟商品'
                    and t.showLocation like '%售后%'
                    and t.date_add>= '{start_date}'
                    AND t.date_add<= '{end_date}'+' 23:59:59'
                    GROUP BY
                        isnull(cv.会员号,t.tel),t.product_title) as q
                JOIN
                    BI_Business_Consume as c
                ON
                    c.会员号 = q.会员号
                JOIN
                    BI_Business_Member as m
                ON	
                    c.会员号 = m.会员号
                AND
                    q.最先兑换时间<=c.消费时间
                GROUP BY
                    m.等级,
                    q.券名
        '''
        cursor.execute(sql3)
        # 写一个如果shquan2是空的话，就不要decode_columns的容错机制
        shquan2 = pd.DataFrame(cursor.fetchall()) 
        if  len(shquan2) > 0:
            shquan2 = decode_columns(shquan2,["等级"])
        else:
            shquan2 = pd.DataFrame(columns=["等级","券名","会员数","转化金额"])

# shquan2
# 读取 Excel 数据并筛选需要的字段
df = shquan1
# 计算不同等级的兑换量和使用量
df_level_sum = df.groupby(['券名', '等级'], as_index=False)['数量'].sum()
df_level_used = df[df['门店受理时间'].notnull()].groupby(['券名', '等级'], as_index=False)['数量'].sum()
# 获取所有的券名转化成一个列表

data = {
    '券名': ['100元售后现金抵用券', '20元售后现金抵用券', '50元售后现金抵用券', '刻字券（售后）', '改圈券（售后）','孔雀石专用券（售后）','编绳券（售后）','编绳券A（售后）','编绳券B（售后）'],
    '彩金': [None, None, None, None, None,None,None,None,None],
    '白金': [None, None, None, None, None,None,None,None,None],
    '紫金': [None, None, None, None, None,None,None,None,None],
    '黑金': [None, None, None, None, None,None,None,None,None],
    '总计': [None, None, None, None, None,None,None,None,None],
    '兑换率': [None, None, None, None, None,None,None,None,None],
    '使用情况_彩金': [None, None, None, None, None,None,None,None,None],
    '使用情况_白金': [None, None, None, None, None,None,None,None,None],
    '使用情况_紫金': [None, None, None, None, None,None,None,None,None],
    '使用情况_黑金': [None, None, None, None, None,None,None,None,None],
    '使用情况_总计': [None, None, None, None, None,None,None,None,None],
    '使用率': [None, None, None, None, None,None,None,None,None],
    '转化人数_彩金': [None, None, None, None, None,None,None,None,None],
    '转化人数_白金': [None, None, None, None, None,None,None,None,None],
    '转化人数_紫金': [None, None, None, None, None,None,None,None,None],
    '转化人数_黑金': [None, None, None, None, None,None,None,None,None],
    '转化人数_总计': [None, None, None, None, None,None,None,None,None],
    '转化率': [None, None, None, None, None,None,None,None,None],
    '转化金额_彩金': [None, None, None, None, None,None,None,None,None],
    '转化金额_白金': [None, None, None, None, None,None,None,None,None],
    '转化金额_紫金': [None, None, None, None, None,None,None,None,None],
    '转化金额_黑金': [None, None, None, None, None,None,None,None,None],
    '转化金额_总计': [None, None, None, None, None,None,None,None,None],
}
df = pd.DataFrame(data)
# 使用 pivot_table 将数据转换为所需格式
pivoted_data = df_level_sum.pivot_table(values='数量', index='券名', columns='等级', fill_value=None, aggfunc='sum')
# 将行索引重置
pivoted_data.reset_index(inplace=True)
# 在原始 DataFrame 中添加新的列
for col in pivoted_data.columns:
    if col not in df.columns and col != '券名':
        df[col] = None
# 将新数据添加到原始 DataFrame 中
for index, row in pivoted_data.iterrows():
    for col in row.index:
        if col not in df.columns and col != '券名':
            continue
        df.loc[df['券名'] == row['券名'], col] = row[col]
# 使用 pivot_table 将数据转换为所需格式
usage_pivoted_data = df_level_used.pivot_table(values='数量', index='券名', columns='等级', fill_value=None, aggfunc='sum')
# 将行索引重置
usage_pivoted_data.reset_index(inplace=True)
# 在原始 DataFrame 中添加新的列，为新数据添加后缀 "_使用"
for col in usage_pivoted_data.columns:
    new_col = f"使用情况_{col}" if col != '券名' else col
    if new_col not in df.columns:
        df[new_col] = None
# 将新数据添加到原始 DataFrame 中
for index, row in usage_pivoted_data.iterrows():
    for col in row.index:
        if col == '券名':
            continue
        new_col = f"使用情况_{col}"
        df.loc[df['券名'] == row['券名'], new_col] = row[col]
membership_data_df = shquan2[["等级","券名","会员数"]]
# 使用 pivot_table 将数据转换为所需格式
pivoted_membership_data = membership_data_df.pivot_table(values='会员数', index='券名', columns='等级', fill_value=None, aggfunc='sum')

# 将行索引重置
pivoted_membership_data.reset_index(inplace=True)

# 在原始 DataFrame 中添加新的列
for col in pivoted_membership_data.columns:
    new_col = f"转化人数_{col}" if col != '券名' else col
    if new_col not in df.columns:
        df[new_col] = None

# 将新数据添加到原始 DataFrame 中
for index, row in pivoted_membership_data.iterrows():
    for col in row.index:
        if col == '券名':
            continue
        new_col = f"转化人数_{col}"
        df.loc[df['券名'] == row['券名'], new_col] = row[col]

money_data_df = shquan2[["等级","券名","转化金额"]]

# 使用 pivot_table 将数据转换为所需格式
pivoted_money_data = money_data_df.pivot_table(values='转化金额', index='券名', columns='等级', fill_value=None, aggfunc='sum')

# 将行索引重置
pivoted_money_data.reset_index(inplace=True)

# 在原始 DataFrame 中添加新的列
for col in pivoted_money_data.columns:
    new_col = f"转化金额_{col}" if col != '券名' else col
    if new_col not in df.columns:
        df[new_col] = None

# 将新数据添加到原始 DataFrame 中
for index, row in pivoted_money_data.iterrows():
    for col in row.index:
        if col == '券名':
            continue
        new_col = f"转化金额_{col}"
        df.loc[df['券名'] == row['券名'], new_col] = row[col]

# 将数据转换为所需格式
df.fillna(0, inplace=True)
# 计算总计列的汇总值
df['总计'] = df[['彩金', '白金', '紫金', '黑金']].sum(axis=1)
# 计算使用情况_总计列的汇总值
df['使用情况_总计'] = df[['使用情况_彩金', '使用情况_白金', '使用情况_紫金', '使用情况_黑金']].sum(axis=1)
# 计算转化人数_总计列的汇总值
df['转化人数_总计'] = df[['转化人数_彩金', '转化人数_白金', '转化人数_紫金', '转化人数_黑金']].sum(axis=1)
# 计算转化金额_总计列的汇总值
df['转化金额_总计'] = df[['转化金额_彩金', '转化金额_白金', '转化金额_紫金', '转化金额_黑金']].sum(axis=1)
# 计算每种券的兑换总张数
df['兑换总张数'] = df[['彩金', '白金', '紫金', '黑金']].sum(axis=1)
# 计算所有兑换优惠券的数量
total_coupons = df['兑换总张数'].sum()
# 计算每种券的兑换率
df['兑换率'] = df['兑换总张数'] / total_coupons
# 计算每种券的使用率
df['使用率'] = np.where(df['兑换总张数'] != 0, df['使用情况_总计'] / df['兑换总张数'], 0)
# 计算每种券的转化率
df['转化人数_总计'] = df['转化人数_总计'].apply(lambda x: float(x))
df['兑换总张数'] = df['兑换总张数'].apply(lambda x: float(x)) 
df['总计'] = df['总计'].apply(lambda x: float(x)) 
df['使用情况_总计'] = df['使用情况_总计'].apply(lambda x: float(x)) 
df['转化率'] = df['转化人数_总计'] / df['兑换总张数']
summary_data = {
    '券名': '汇总',
    '彩金': df['彩金'].sum(),
    '白金': df['白金'].sum(),
    '紫金': df['紫金'].sum(),
    '黑金': df['黑金'].sum(),
    '总计': df['总计'].sum(),
    '兑换率':1,
    '使用情况_彩金': df['使用情况_彩金'].sum(),
    '使用情况_白金': df['使用情况_白金'].sum(),
    '使用情况_紫金': df['使用情况_紫金'].sum(),
    '使用情况_黑金': df['使用情况_黑金'].sum(),
    '使用情况_总计': df['使用情况_总计'].sum(),
    '使用率': df['使用情况_总计'].sum()/ df['总计'].sum() if df['总计'].sum() != 0 else 0,
    '转化人数_彩金': df['转化人数_彩金'].sum(),
    '转化人数_白金': df['转化人数_白金'].sum(),
    '转化人数_紫金': df['转化人数_紫金'].sum(),
    '转化人数_黑金': df['转化人数_黑金'].sum(),
    '转化人数_总计': df['转化人数_总计'].sum(),
    '转化率': df['转化人数_总计'].sum()/df['总计'].sum() if df['总计'].sum() != 0 else 0,
    '转化金额_彩金': df['转化金额_彩金'].sum(),
    '转化金额_白金': df['转化金额_白金'].sum(),
    '转化金额_紫金': df['转化金额_紫金'].sum(),
    '转化金额_黑金': df['转化金额_黑金'].sum(),
    '转化金额_总计': df['转化金额_总计'].sum(),
}
df.drop(columns='兑换总张数',inplace=True)
df[['兑换率','使用率','转化率']]=df[['兑换率','使用率','转化率']].applymap(lambda x: '{:.2%}'.format(x))
df.fillna("-",inplace=True)
shquan = df
# 读取 Excel 数据并筛选需要的字段
with pymssql.connect(server, user, password,database) as conn:
    with conn.cursor(as_dict=True) as cursor:
        # 兑礼数据
        sql4 = f'''
                    SELECT
                        t1.[等级],
                        t1.[兑换数],
                        t1.[兑换人数],
                        t2.转化人数,
                        t2.[转化金额]
                    FROM
                    (SELECT
                        m.等级,
                        count(v.no) as 兑换数,
                        count(distinct v.tel) as 兑换人数
                    FROM
                        v_mall_order_detail as v
                    JOIN
                        bi_business_member as m
                    ON
                        v.tel = m.会员号
                    where 
                        date_add>='{start_date}'
                    AND
                        date_add<='{end_date}'
                    AND
                        sku = 'QJ00000000700001'
                    AND 
                        statuname != '关闭' 
                    AND statuname != '申请退货' 
                    AND statuname != '已退款' 
                    GROUP BY
                        m.等级) as t1
                    JOIN
                    (SELECT 
                                c.等级,
                                count(distinct c.会员号) as 转化人数,
                                SUM(c.金额) AS 转化金额
                        FROM 
                                (
                                        SELECT 
                                                v.tel, 
                                                MIN(v.date_add) AS 最新兑换记录 
                                        FROM 
                                                v_mall_order_detail AS v 
                                                INNER JOIN (
                                                        SELECT 
                                                                tel, 
                                                                MAX(date_add) AS date_add 
                                                        FROM 
                                                                v_mall_order_detail 
                                                        WHERE 
                                                                date_add BETWEEN '{start_date}' AND '{end_date}'
                                                                AND sku = 'QJ00000000700001' 
                                                                AND statuname NOT IN ('关闭', '申请退货', '已退款') 
                                                        GROUP BY 
                                                                tel
                                                ) AS sub 
                                                ON v.tel = sub.tel AND v.date_add = sub.date_add 
                                        GROUP BY 
                                                v.tel
                                ) AS t1 
                                INNER JOIN (
                                        SELECT 
                                                m.会员号, 
                                                m.等级, 
                                                c.消费时间, 
                                                c.金额 
                                        FROM 
                                                BI_Business_Member AS m 
                                                INNER JOIN bi_business_consume AS c 
                                                        ON m.会员号 = c.会员号 
                                        WHERE 
                                                c.消费时间 BETWEEN '{start_date}' AND '{end_date}'
                                ) AS c 
                                        ON t1.tel = c.会员号 AND c.消费时间 >= t1.最新兑换记录 
                        GROUP BY 
                                c.等级) as t2 
                    on 
                        t1.等级 = t2.等级
                '''
        cursor.execute(sql4)
        # 创建一个空的dataframe
        df = pd.DataFrame(columns=['等级','兑换数','兑换人数','转化人数','转化金额'])
        df.loc[0] = ['彩金',0,0,0,0] 
        df.loc[1] = ['白金',0,0,0,0]
        df.loc[2] = ['紫金',0,0,0,0] # type: ignore
        df.loc[3] = ['黑金',0,0,0,0]
        pisheng = pd.DataFrame(cursor.fetchall())
        # 怕段pisheng是否为空
        if pisheng.empty:
                pisheng = df
        else:
        # 将pisheng各个等级的数据填充到df中
            pisheng = decode_columns(pisheng,["等级"])
            for i in range(len(pisheng)):
                    df.loc[i] = pisheng.loc[i]
        
        # 计算转化率
        # 判断是否有兑换人数为0的情况
        try:
            df["转化率"] = df["转化人数"]/df["兑换人数"]
        except:
            df["转化率"] = 0 

        df.fillna('-',inplace=True)
        order = ['等级', '兑换数', '兑换人数', '转化人数','转化率','转化金额']
        df= df[order]
        pisheng = df
from openpyxl import load_workbook
# 读取数据
# 假设你的 DataFrame 名称为 result2
# result2 = pd.DataFrame(...)
# 加载现有的 Excel 文件
file_path = r'.\消费7天后回访触达情况.xlsx'
book = load_workbook(file_path)

# 选择要将数据写入的工作表
sheet_name1 = '大区'  # 请更改为你的工作表名称
sheet1 = book[sheet_name1]
# 将 DataFrame 数据写入工作表的指定位置
start_row = 4  # 起始行
start_col = 1  # 起始列
for index, row in enumerate(result2_final.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet1.cell(row=start_row + index, column=start_col + j, value=value)
# 选择要将数据写入的工作表
sheet_name2 = '小区'  # 请更改为你的工作表名称
sheet2 = book[sheet_name2]
# 将 DataFrame 数据写入工作表的指定位置
start_row = 4  # 起始行
start_col = 1  # 起始列
for index, row in enumerate(result1_final.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet2.cell(row=start_row + index, column=start_col + j, value=value)
# 选择要将数据写入的工作表
sheet_name3 = '外拨任务效果'  # 请更改为你的工作表名称
sheet3 = book[sheet_name3]
# 将 DataFrame 数据写入工作表的指定位置
start_row = 2  # 起始行
start_col = 1  # 起始列
for index, row in enumerate(chuda.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet3.cell(row=start_row + index, column=start_col + j, value=value)

# 选择要将数据写入的工作表
sheet_name4 = '售后券'  # 请更改为你的工作表名称
sheet4 = book[sheet_name4]

# 将 DataFrame 数据写入工作表的指定位置
start_row = 3  # 起始行
start_col = 1  # 起始列

for index, row in enumerate(shquan.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet4.cell(row=start_row + index, column=start_col + j, value=value)    

# 选择要将数据写入的工作表
sheet_name5 = '皮绳'  # 请更改为你的工作表名称
sheet5 = book[sheet_name5]
# 将 DataFrame 数据写入工作表的指定位置
start_row = 2  # 起始行
start_col = 1  # 起始列

for index, row in enumerate(df.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet5.cell(row=start_row + index, column=start_col + j, value=value)      
# 保存更改到 Excel 文件
book.save(file_path)
print('写入成功')