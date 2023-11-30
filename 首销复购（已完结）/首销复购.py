# -*- coding: utf-8 -*-
import pandas as pd
import pymssql 
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# 数据处理
def decode_columns(df, column_names):
    """
    对pandas数据框中的多个列进行编解码，并返回新的数据框
    
    Args:
        chuda: pandas数据框
        column_names: list，需要进行编解码的列名列表
    
    Returns:
        pandas数据框，编解码后的结果
    """
    for column_name in column_names:
        # 跳过空值
        mask = df[column_name].notnull()
        # 对指定列进行编解码
        df.loc[mask, column_name] = df.loc[mask, column_name].apply(lambda x: x.encode('latin1').decode('gbk'))
    # 返回编解码后的结果
    return df
# 按照周份
def get_dynamic_dates_weekly():
    # 获取今天的日期
    today = datetime.date.today()
    # 获取今天是周几（0代表周一，1代表周二，依次类推）
    weekday = today.weekday()
    if weekday == 0: # 今天是周一
        # 开始时间设置为上周周一
        start_date = today - datetime.timedelta(days=7)
        # 结束时间设置为上周周末
        end_date = today - datetime.timedelta(days=1)
    else: # 今天不是周一
        # 开始时间设置为这周周一
        start_date = today - datetime.timedelta(days=16)
        # 结束时间设置为昨天
        end_date = today - datetime.timedelta(days=1)
    # 返回开始时间和结束时间
    return start_date, end_date
# 按照月份
def get_dynamic_dates():
    # 获取今天的日期
    today = datetime.date.today()
    # 获取本月的第一天
    first_day_of_month = today.replace(day=1)
    # 判断今天是否在本月的3天内
    if today.day <4:
        # 如果是，将开始时间设置为上个月的1号
        start_date = (first_day_of_month - datetime.timedelta(days=1)).replace(day=1)
    else:
        # 如果不是，将开始时间设置为本月的1号
        start_date = first_day_of_month
    
    # 将结束时间设置为昨天
    end_date = today - datetime.timedelta(days=2)
    
    # 返回开始时间和结束时间
    return start_date, end_date
# start_date, end_date = get_dynamic_dates_weekly()
start_date, end_date = get_dynamic_dates()
params = {"start_date": start_date, "end_date": end_date}
round_func = lambda x: round(x, 2)
print(params["start_date"])
print(params["end_date"])
server = '192.168.0.169'
user = 'chjreport'
password = 'Chj@12345'
database = 'ChjBidb'
with pymssql.connect(server, user, password,database) as conn:
    with conn.cursor(as_dict=True) as cursor:
        sql1 = f'''
                with t as (--七天回访明细
                select v.大区名称 大区,v.小区名称 小区,v.门店代码,v.门店名称,id,taskno,taskname,status,memberid,membercode,startTime,
                                case when t.taskName like '%消费7天%' then '七天回访' when t.taskName like '%复购精准%' then '30天首销回访' else null end 回访类型
                from v_crm_shoptask_detail t
                left join v_shop_dict v on v.门店代码=t.revisitshopcode
                where (t.taskName like '%消费7天%')
                and t.startTime between '2023-06-01' and '2023-06-30 23:59:59'
                and t.revisitGuideNo not in (select distinct guide_no from v_mall_guide_all where status='5')
                and v.大区名称='西北大区'
                )
                ,sx as (--新老会员：是否为首销会员（近1年消费1次且为首销，包括沉睡，流失会员重新回柜的用户）
                select cc.会员ID
                from t
                join BI_Business_Consume cc on cc.会员ID=t.memberid AND 消费时间 between '2022-04-11' and '2023-06-30 23:59:59.99' and 金额>0 
                group by cc.会员ID    
                having count(distinct 单据编号)=1 and count(distinct case when 消费时间 between '2022-06-11' and '2023-06-30 23:59:59.99' then 单据编号 else null end)=1
                )
                ,hy as (--剔除活跃会员
                        select t.会员ID
                        from (
                                select cc.会员ID
                                from t
                                join BI_Business_Consume cc on cc.会员ID=t.memberid AND 消费时间 between '2022-04-11' and '2023-06-30 23:59:59.99' and 金额>0 
                                group by cc.会员ID    
                                having count(distinct 单据编号)>=1
                        )t
                        join BI_Business_Member cv  on t.会员ID=cv.会员ID 
                        where cv.共消费次数 > 1 
                )
                ,tt as (--新老会员
                select t.*,case when 回访类型='七天回访' and sx.会员ID is null then '老会员' else '新会员' end 新老会员
                from t
                left join sx on t.memberid=sx.会员ID
                left join hy on t.memberid=hy.会员ID
                where hy.会员ID is null
                )
                ,q1 as (--添加企微的会员
                select distinct 门店代码,t.会员号
                from BI_Business_qywechatmember_all  t
                where t.会员号 is not  null
                )
                ,q2 as (--贴标会员
                select distinct cv.会员ID,g.store_no 门店代码--,tel 导购手机号  --,CV.会员号
                from BI_Business_qywechatmember_tag t
                join BI_Business_Member cv on t.unionid=cv.微信unionId
                left join v_mall_guide g on g.userid=t.qyuserid
                )
                ,fg as (--复购
                select 回访类型,t.大区,t.小区,t.门店代码,t.门店名称,新老会员,count(distinct t.memberid) 复购人数,sum(cc.金额)/10000 转化金额
                from BI_Business_Consume cc
                join (--统计期内最早一次回访时间
                                select 回访类型,大区,小区,门店代码,t.门店名称,新老会员,memberid,cv.第一次消费时间,min(startTime) startTime
                                from tt t
                                join BI_Business_Member cv on cv.会员ID=t.memberid
                                where status=1
                                group by 回访类型,大区,小区,门店代码,t.门店名称,新老会员,memberid,cv.第一次消费时间
                                )t on t.memberId=cc.会员ID AND CC.消费时间>t.startTime and CC.消费时间<>t.第一次消费时间
                join v_btgoods as g on cc.商品代码 = g.商品代码 
                where cc.消费时间 between '2023-06-01' and '2023-06-30 23:59:59.99'  and t.门店代码=cc.门店代码 and 二级大类 != '促销物料'
                group by 回访类型,t.大区,t.小区,t.门店代码,t.门店名称,新老会员
                )
                select t.*,fg.复购人数,fg.转化金额
                from (
                                select 回访类型,大区,小区,t.门店代码,门店名称,t.新老会员,
                                                count(1) 任务数,sum(case when t.status='1' then 1 else 0 end) 完成任务数,
                                                cast(cast((sum(case when t.status='1' then 1 else 0 end)*100.0/count(1)) as  numeric(10,2)) as varchar) 触达完成率,
                                                sum(case when q1.会员号 is not null then 1 else 0 end) 企微添加数,
                                                cast(cast((sum(case when q1.会员号 is not null then 1 else 0 end)*100.0/count(1) ) as  numeric(10,2)) as varchar) 企微添加率,
                                                sum(case when q2.会员ID is not null then 1 else 0 end) 贴标人数,
                                                cast(cast((sum(case when q2.会员ID is not null then 1 else 0 end)*100.0/count(1) ) as  numeric(10,2)) as varchar) 贴标率
                                from tt t
                                left join q1 on t.memberCode=q1.会员号 and t.门店代码=q1.门店代码
                                left join q2 on t.memberid=q2.会员ID and t.门店代码=q2.门店代码
                                group by 回访类型,大区,小区,t.门店代码,门店名称,t.新老会员
                )t
                left join fg on t.回访类型=fg.回访类型 and t.门店代码=fg.门店代码 and t.新老会员=fg.新老会员 
                order by t.小区,t.门店代码,t.新老会员

                '''
        sql2 = f'''
                with t as (--七天回访明细
                select v.大区名称 大区,v.小区名称 小区,v.门店代码,v.门店名称,id,taskno,taskname,status,memberid,membercode,startTime,
                                case when t.taskName like '%消费7天%' then '七天回访' when t.taskName like '%复购精准%' then '30天首销回访' else null end 回访类型
                from v_crm_shoptask_detail t
                left join v_shop_dict v on v.门店代码=t.revisitshopcode
                where (t.taskName like '%复购精准%')
                and t.startTime between '2023-06-01' and '2023-06-30 23:59:59.99' 
                and t.revisitGuideNo not in (select distinct guide_no from v_mall_guide_all where status='5')
                and v.大区名称='西北大区'
                )
                ,q1 as (--添加企微的会员
                select distinct 门店代码,t.会员号
                from BI_Business_qywechatmember_all  t
                where t.会员号 is not  null
                )
                ,q2 as (--贴标会员
                select distinct cv.会员ID,g.store_no 门店代码--,tel 导购手机号  --,CV.会员号
                from BI_Business_qywechatmember_tag t
                join BI_Business_Member cv on t.unionid=cv.微信unionId
                left join v_mall_guide g on g.userid=t.qyuserid
                )
                ,fg as (--复购
                select 回访类型,t.大区,t.小区,t.门店代码,t.门店名称,count(distinct t.memberid) 复购人数,sum(cc.金额)/10000 转化金额
                from BI_Business_Consume cc
                join (--统计期内最早一次回访时间
                                select 回访类型,大区,小区,门店代码,t.门店名称,memberid,cv.第一次消费时间,min(startTime) startTime
                                from t
                                join BI_Business_Member cv on cv.会员ID=t.memberid
                                where status=1
                                group by 回访类型,大区,小区,门店代码,t.门店名称,memberid,cv.第一次消费时间
                                )t on t.memberId=cc.会员ID AND CC.消费时间>t.startTime and CC.消费时间<>t.第一次消费时间
                where cc.消费时间 between '2023-06-01' and '2023-06-30 23:59:59.99'  and t.门店代码=cc.门店代码 
                group by 回访类型,t.大区,t.小区,t.门店代码,t.门店名称
                )
                select t.*,fg.复购人数,fg.转化金额
                from (
                                select 回访类型,大区,小区,t.门店代码,门店名称,
                                                count(1) 任务数,sum(case when t.status='1' then 1 else 0 end) 完成任务数,
                                                cast(cast((sum(case when t.status='1' then 1 else 0 end)*100.0/count(1)) as  numeric(10,2)) as varchar) 触达完成率,
                                                sum(case when q1.会员号 is not null then 1 else 0 end) 企微添加数,
                                                cast(cast((sum(case when q1.会员号 is not null then 1 else 0 end)*100.0/count(1) ) as  numeric(10,2)) as varchar) 企微添加率,
                                                sum(case when q2.会员ID is not null then 1 else 0 end) 贴标人数,
                                                cast(cast((sum(case when q2.会员ID is not null then 1 else 0 end)*100.0/count(1) ) as  numeric(10,2)) as varchar) 贴标率
                                from t
                                left join q1 on t.memberCode=q1.会员号 and t.门店代码=q1.门店代码
                                left join q2 on t.memberid=q2.会员ID and t.门店代码=q2.门店代码
                                group by 回访类型,大区,小区,t.门店代码,门店名称
                )t
                left join fg on t.回访类型=fg.回访类型 and t.门店代码=fg.门店代码
                order by t.小区,t.门店代码
                '''

        # 7天回访：“消费7天”
        cursor.execute(sql1)
        chuda1 = pd.DataFrame(cursor.fetchall())
        chuda1 = decode_columns(chuda1,["回访类型","新老会员"])
        # 30天首销回访：“复购精准”
        cursor.execute(sql2)
        chuda2 = pd.DataFrame(cursor.fetchall())
        chuda2 = decode_columns(chuda2,["回访类型"])
        # 把chuda1和chuda2的数字列都转成数字类型
        for i in chuda1.columns:
                if i not in ["回访类型","大区","小区","门店代码","门店名称","新老会员"]:
                        chuda1[i] = pd.to_numeric(chuda1[i])
        for i in chuda2.columns:
                if i not in ["回访类型","大区","小区","门店代码","门店名称"]:
                        chuda2[i] = pd.to_numeric(chuda2[i])
server = '192.168.0.188'
user = 'YFE'
password = 'Chj@2021'
database = 'MemberCenter'
with pymssql.connect(server, user, password,database) as conn:
    with conn.cursor(as_dict=True) as cursor:
        sql1 = f'''
                        with t1 as (--应发会员数
                        SELECT distinct LargeAreaName,CommunityName,revisitshopcode,ShopName,memberCode,status
                        FROM [dbo].[Business_ShopTask_Detail] a 
                        left join [Business_ShopTask] b on a.taskid=b.id
                        left join   (SELECT  ShopCode, MAX(ShopName) AS ShopName, MAX(LargeAreaName) AS LargeAreaName, 
                                                                MAX(CommunityName) AS CommunityName, MAX(shopProperty) AS shopProperty
                                                FROM       dbo.Base_Shop
                                                GROUP BY ShopCode) AS shop on a.revisitshopcode = shop.ShopCode 
                        where b.taskname like '%30天回访%' and startTime between '2023-06-01' and '2023-06-30'+' 23:59:59'
                        and shop.LargeAreaName='西北大区'
                        )
                        ,t2 as (---实际完成任务
                        SELECT distinct m3.membercode,m3.FirstConsumeTime,g.departmentcode
                        FROM  [dbo].[Business_qywechat_mass_record] AS r 
                        JOIN Business_qywechat_mass_record_member AS m1 ON r.msgid = m1.msgid
                        JOIN Business_qywechatmember AS m2 ON m1.external_userid = m2.externalUserId
                        JOIN Business_Member AS m3 ON m2.unionId = m3.wxUnionId and brandid='A637A7F84DF54960AE658745FCB8208F'
                        join t1  on t1.memberCode=m3.memberCode
                        left join  base_staff g on m1.userid=g.userid   --userid,mobile,departmentcode
                        WHERE m1.task_status=2 and (content like '%特邀会员的专属福利%') and r.create_time between '2023-06-01' and '2023-06-30'+' 23:59:59'
                        )
                        ,t3 as (--点击链接
                        select t.mobile,t2.FirstConsumeTime,min(t.date_add) date_add
                        from (
                        select 
                        t.id,t.openid,t.unionid,t.member_id,t.url,t.date_add,t.options,t.prodid,t.type,t.sessionId,t.mobile,
                                        CASE WHEN [options] LIKE '%RecoID":"%' THEN replace(SUBSTRING([options], CHARINDEX('RecoID":"', [options]) + 9, 
                                        CHARINDEX('"', [options], CHARINDEX('RecoID":"', [options]) + 9) - (CHARINDEX('RecoID":"', [options]) + 9)), '\n', '') 
                                        WHEN [options] LIKE '%RecoID=%' THEN SUBSTRING([options], CHARINDEX('RecoID=', [options]) + 7, CHARINDEX('"', 
                                        [options], CHARINDEX('RecoID=', [options]) + 7) - (CHARINDEX('RecoID=', [options]) + 7)) 
                                        WHEN [options] LIKE '%RecoID%3D%' THEN SUBSTRING([options], CHARINDEX('RecoID%3D', [options]) + 9, 
                                        CHARINDEX('"', [options], CHARINDEX('RecoID%3D', [options]) + 9) - (CHARINDEX('RecoID%3D', [options]) + 9)) ELSE NULL 
                                        END AS sourceMemberId
                        from DigitalRetailDB.dbo.Business_Visit_Record t
                        where (t.url='pages/landing/gift/list')
                        and date_add between '2023-06-01' and '2023-06-30'+' 23:59:59'
                        and  (options LIKE '%RecoID%') 
                        )t
                        join t2 on t2.membercode=t.mobile
                        where sourceMemberId=1614 and t.mobile is not null
                        group by t.mobile,t2.FirstConsumeTime
                        )
                        ,fg as (
                        select 门店代码,sum(转化金额) 转化金额, count(distinct t.mobile) 复购人数
                        from  (
                                select  cc.门店代码,t3.mobile,sum(cc.金额)/10000 转化金额
                                from t3 
                                join BI_Business_Consume cc on t3.mobile=cc.会员号 and cc.消费时间<>t3.FirstConsumeTime and cc.消费时间>t3.date_add
                                where cc.消费时间 between '2023-06-01' and '2023-06-30'+' 23:59:59'
                                group by cc.门店代码,t3.mobile
                                )t
                        group by 门店代码
                        )
                        select t.大区,t.小区,t.门店代码,t.门店名称,t.工单任务数,t.实际完成任务数量,
                                                        case when t.工单任务数=0  then '0'  else cast(cast(t.实际完成任务数量*100.0/t.工单任务数 as  numeric(10,2)) as varchar)+'%' end 触达完成率,
                                                        点击链接数,
                                                        case when t.实际完成任务数量=0 then '0' else cast(cast(t.点击链接数*100.0/t.实际完成任务数量 as  numeric(10,2)) as varchar)+'%' end 点击率,
                                                        fg.复购人数,fg.转化金额
                        from (
                        select LargeAreaName 大区,CommunityName 小区,revisitshopcode 门店代码,ShopName 门店名称,
                                                count(distinct t1.memberCode) 工单任务数,--count(distinct case when t1.status =1 then t1.memberCode else null end) 实际完成任务数,---,min()
                        count(distinct case when t2.memberCode is not null then t2.memberCode else null end) 实际完成任务数量 ,
                        count(distinct case when t3.mobile is not null then t3.mobile else null end) 点击链接数
                        from t1
                        left join t2 on t1.memberCode=t2.memberCode and t1.revisitshopcode=t2.departmentcode
                        left join t3 on t1.memberCode=t3.mobile
                        group by LargeAreaName,CommunityName,revisitshopcode,ShopName
                        )t
                        left join fg on  t.门店代码=fg.门店代码
                '''
        # 30天：“一键触达”
        cursor.execute(sql1)
        chuda3 = pd.DataFrame(cursor.fetchall())
        if chuda3.empty:
        #生成空数据
                chuda3 = pd.DataFrame(columns=['大区','小区','门店代码','门店名称','工单任务数','实际完成任务数量','触达完成率','点击链接数','点击率','复购人数','转化金额'])
        else:
            chuda3 = decode_columns(chuda3,["大区","小区","门店名称"])
        # 帮吧把触达完成率和点击率的百分号去掉，并转化类型为数字
        chuda3['触达完成率'] = chuda3['触达完成率'].str.replace('%', '').astype(float)
        chuda3['点击率'] = chuda3['点击率'].str.replace('%', '').astype(float)

# 按小区分组并计算统计数据
grouped_30_days = chuda2.groupby('小区').agg({
    '任务数': 'sum',
    '完成任务数': 'sum',
    '触达完成率': 'mean',
    '企微添加数': 'sum',
    '企微添加率': 'mean',
    '贴标人数': 'sum',
    '贴标率': 'mean',
    '复购人数': 'sum',
    '转化金额': 'sum'
})
grouped_7_days = chuda1.groupby(['小区']).agg({
    '任务数': 'sum',
    '完成任务数': 'sum',
    '触达完成率': 'mean',
    '企微添加数': 'sum',
    '企微添加率': 'mean',
    '贴标人数': 'sum',
    '贴标率': 'mean',
    '复购人数': 'sum',
    '转化金额': 'sum'
})
grouped_yijian = chuda3.groupby(['小区']).agg({
    '工单任务数': 'sum',
    '实际完成任务数量': 'sum',
    '触达完成率': 'mean',
    '点击链接数': 'sum',
    '点击率': 'mean',
    '复购人数': 'sum',
    '转化金额': 'sum'
})
grouped_7_days['触达完成率'] = grouped_7_days['完成任务数']/grouped_7_days['任务数']
grouped_7_days['企微添加率'] = grouped_7_days['企微添加数']/grouped_7_days['任务数']
grouped_7_days['贴标率'] = grouped_7_days['贴标人数']/grouped_7_days['企微添加数']
grouped_30_days['触达完成率'] = grouped_30_days['完成任务数']/grouped_30_days['任务数']
grouped_30_days['企微添加率'] = grouped_30_days['企微添加数']/grouped_30_days['任务数']
grouped_30_days['贴标率'] = grouped_30_days['贴标人数']/grouped_30_days['企微添加数']
# 帮我为grouped_30_days和grouped_7_days的列名改成多级索引
grouped_30_days.columns = pd.MultiIndex.from_product([['30天'], grouped_30_days.columns])
grouped_7_days.columns = pd.MultiIndex.from_product([['7天'], grouped_7_days.columns])
grouped_yijian.columns = pd.MultiIndex.from_product([['一键触达'], grouped_yijian.columns])
# 将三个grouped合并，用merge
grouped = pd.merge( grouped_7_days, grouped_30_days,left_index=True, right_index=True, how='outer')
grouped = pd.merge(grouped, grouped_yijian, left_index=True, right_index=True, how='outer')
result_final = grouped
result_final['7天', '触达完成率'] = (result_final['7天', '触达完成率']).apply(lambda x: format(x, '.2%'))
result_final['7天', '贴标率'] = (result_final['7天', '贴标率']/100).apply(lambda x: format(x, '.2%'))
result_final['7天', '企微添加率'] = (result_final['7天', '企微添加率']).apply(lambda x: format(x, '.2%'))
result_final['30天', '触达完成率'] = (result_final['30天', '触达完成率']).apply(lambda x: format(x, '.2%'))
result_final['30天', '贴标率'] = (result_final['30天', '贴标率']/100).apply(lambda x: format(x, '.2%'))
result_final['30天', '企微添加率'] = (result_final['30天', '企微添加率']).apply(lambda x: format(x, '.2%'))
result_final['一键触达', '触达完成率'] = (result_final['一键触达', '触达完成率']/100).apply(lambda x: format(x, '.2%'))
result_final['一键触达','点击率'] = result_final['一键触达','点击链接数']/result_final['一键触达','实际完成任务数量']
result_final['一键触达', '点击率'] = (result_final['一键触达', '点击率']).apply(lambda x: format(x, '.2%'))
# 去除所有贴标人数和贴标率的列
result_final2 = result_final.drop(columns=[('7天', '贴标人数'), ('7天', '贴标率'), ('30天', '贴标人数'), ('30天', '贴标率')])
# 把字符串为nan%的数据替换成0%
result_final2 = result_final2.replace('nan%', '0%')
result_final2.fillna(0, inplace=True)

shop_7_days = chuda1
shop_30_days = chuda2
shop_yijian = chuda3
# # chuda1只要新老会员里面的新会员
# shop_7_days = shop_7_days[shop_7_days['新老会员'] == '新会员']
# # # 删除chuda1中的新老会员列
shop_7_days = shop_7_days.drop(columns=['新老会员'])
# # 删除回访类型列
shop_7_days = shop_7_days.drop(columns=['回访类型'])
# 删除chuda2中的回访类型
shop_30_days = shop_30_days.drop(columns=['回访类型'])
# 把nan替换成0
shop_7_days = shop_7_days.fillna(0)
shop_30_days = shop_30_days.fillna(0)
shop_yijian = shop_yijian.fillna(0)
shop_7_days = shop_7_days.set_index(['大区', '小区', '门店代码', '门店名称'])
shop_30_days = shop_30_days.set_index(['大区', '小区', '门店代码', '门店名称'])
shop_yijian = shop_yijian.set_index(['大区', '小区', '门店代码', '门店名称'])
# 帮我为grouped_30_days和grouped_7_days的列名改成多级索引
shop_7_days.columns = pd.MultiIndex.from_product([['7天'], shop_7_days.columns])
shop_30_days.columns = pd.MultiIndex.from_product([['30天'], shop_30_days.columns])
shop_yijian.columns = pd.MultiIndex.from_product([['一键触达'], shop_yijian.columns])
# 将三个grouped合并，用merge
shop = pd.merge( shop_7_days, shop_30_days,left_index=True, right_index=True, how='outer')
shop = pd.merge(shop, shop_yijian, left_index=True, right_index=True, how='outer')
# 删除贴标相关的列
shop = shop.drop(columns=[('7天', '贴标人数'), ('7天', '贴标率'), ('30天', '贴标人数'), ('30天', '贴标率')])
shop_final = shop
shop_final['7天', '触达完成率'] = (shop_final['7天', '触达完成率']/100).apply(lambda x: format(x, '.2%'))
shop_final['7天', '企微添加率'] = (shop_final['7天', '企微添加率']/100).apply(lambda x: format(x, '.2%'))
shop_final['30天', '触达完成率'] = (shop_final['30天', '触达完成率']/100).apply(lambda x: format(x, '.2%'))
shop_final['30天', '企微添加率'] = (shop_final['30天', '企微添加率']/100).apply(lambda x: format(x, '.2%'))
shop_final['一键触达', '触达完成率'] = (shop_final['一键触达', '触达完成率']/100).apply(lambda x: format(x, '.2%'))
shop_final['一键触达','点击率'] = shop_final['一键触达','点击链接数']/shop_final['一键触达','实际完成任务数量']
shop_final['一键触达', '点击率'] = (shop_final['一键触达', '点击率']/100).apply(lambda x: format(x, '.2%'))
# 空值替换成0
shop_final = shop_final.fillna(0)
# 把小区，门店代码，门店名称取消索引
shop_final = shop_final.reset_index()
# 替换那些字符串为nan%字符串的为0%
shop_final = shop_final.set_index(['大区', '小区', '门店代码', '门店名称'])
# 解除三个索引
shop_final = shop_final.reset_index()
# 把大区作为索引
shop_final = shop_final.set_index(['大区'])
# 把nan%替换成0%
shop_final = shop_final.replace('nan%', '0%')
shop_final2 = shop_final
# 计算shop_final2的合计行
shop_final2['7天', '触达完成率'] = shop_final2['7天', '触达完成率'].astype(str)
shop_final2['7天', '企微添加率'] = shop_final2['7天', '企微添加率'].astype(str)
shop_final2['30天', '触达完成率'] = shop_final2['30天', '触达完成率'].astype(str)
shop_final2['30天', '企微添加率'] = shop_final2['30天', '企微添加率'].astype(str)
shop_final2['一键触达', '触达完成率'] = shop_final2['一键触达', '触达完成率'].astype(str)
shop_final2['一键触达', '点击率'] = shop_final2['一键触达', '点击率'].astype(str)
shop_final2['7天', '触达完成率'] = shop_final2['7天', '触达完成率'].apply(lambda x: float(x.replace('%', '')))
shop_final2['7天', '企微添加率'] = shop_final2['7天', '企微添加率'].apply(lambda x: float(x.replace('%', '')))
shop_final2['30天', '触达完成率'] = shop_final2['30天', '触达完成率'].apply(lambda x: float(x.replace('%', '')))
shop_final2['30天', '企微添加率'] = shop_final2['30天', '企微添加率'].apply(lambda x: float(x.replace('%', '')))
shop_final2['一键触达', '触达完成率'] = shop_final2['一键触达', '触达完成率'].apply(lambda x: float(x.replace('%', '')))
shop_final2['一键触达', '点击率'] = shop_final2['一键触达', '点击率'].apply(lambda x: float(x.replace('%', '')))
# 计算合计行
shop_final2.loc['合计'] = shop_final2.apply(lambda x: x.sum())
# 计算合计行的触达完成率和企微添加率
shop_final2[('7天', '触达完成率')] = shop_final2[('7天', '完成任务数')]/shop_final2[('7天', '任务数')]
shop_final2[('7天', '企微添加率')] = shop_final2[('7天', '企微添加数')]/shop_final2[('7天', '任务数')]
shop_final2[('30天', '触达完成率')] = shop_final2[('30天', '完成任务数')]/shop_final2[('30天', '任务数')]
shop_final2[('30天', '企微添加率')] = shop_final2[('30天', '企微添加数')]/shop_final2[('30天', '任务数')]
shop_final2[('一键触达', '触达完成率')] = shop_final2[('一键触达', '实际完成任务数量')]/shop_final2[('一键触达', '工单任务数')]
shop_final2[('一键触达', '点击率')] = shop_final2[('一键触达', '点击率')]/shop_final2[('一键触达', '实际完成任务数量')]
# 把整个df的所有率的列转换成百分比
shop_final2['7天', '触达完成率'] = (shop_final2['7天', '触达完成率']).apply(lambda x: format(x, '.2%'))
shop_final2['7天', '企微添加率'] = (shop_final2['7天', '企微添加率']).apply(lambda x: format(x, '.2%'))
shop_final2['30天', '触达完成率'] = (shop_final2['30天', '触达完成率']).apply(lambda x: format(x, '.2%'))
shop_final2['30天', '企微添加率'] = (shop_final2['30天', '企微添加率']).apply(lambda x: format(x, '.2%'))
shop_final2['一键触达', '触达完成率'] = (shop_final2['一键触达', '触达完成率']/100).apply(lambda x: format(x, '.2%'))
shop_final2['一键触达', '点击率'] = (shop_final2['一键触达', '点击率']).apply(lambda x: format(x, '.2%'))
shop_final = shop_final.reset_index()
# 把最后一行的合计行的大区，小区，门店代码，门店名称列的内容替换成合计
# 替换那些字符串为nan%字符串的为0%
# 解除三个索引
shop_final2 = shop_final2.reset_index()
# 把大区作为索引
# shop_final2 = shop_final2.set_index(['大区'])
shop_final2 = shop_final2.replace('nan%', '0%')
shop_final2.loc['合计', '小区'] = '合计'
shop_final2.loc['合计', '门店代码'] = '-'
shop_final2.loc['合计', '门店名称'] = '总计'
# 加载现有的 Excel 文件
file_path = r'D:\code\潮宏基\首销复购（已完结）\西北首销复购数据 - 副本.xlsx'
book = load_workbook(file_path)
# 选择要将数据写入的工作表
sheet_name1 = '汇总'  # 请更改为你的工作表名称
sheet1 = book[sheet_name1]
# 将 DataFrame 数据写入工作表的指定位置
start_row = 4  # 起始行
start_col = 9  # 起始列
for index, row in enumerate(result_final2.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet1.cell(row=start_row + index, column=start_col + j, value=value)
# 选择要将数据写入的工作表
sheet_name2 = '分区'  # 请更改为你的工作表名称
sheet2 = book[sheet_name2]
start_row = 4  # 起始行
start_col = 13  # 起始列
for index, row in enumerate(shop_final2.iterrows()):
    for j, value in enumerate(row[1]):  # 注意这里的变化：使用 row[1] 而不是 row
        sheet2.cell(row=start_row + index, column=start_col + j, value=value)
# 选择要将数据写入的工作表
# 保存更改到 Excel 文件
book.save(file_path)
print("任务完成！！！！")