# -*- coding: utf-8 -*-
import pandas as pd
import pymssql 
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def decode_columns(result, column_names):
    for column_name in column_names:
        # 检查列是否为空
        if result[column_name].notnull().any():
            # 使用Pandas的向量化字符串操作进行编解码
            result[column_name] = result[column_name].str.encode('latin1').str.decode('gbk')
    return result

def get_dynamic_dates_weekly():
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=7 + today.weekday())
    end_date = today - datetime.timedelta(days=today.weekday())
    return start_date, end_date

def get_dynamic_dates():
    today = datetime.date.today()
    first_day_of_month = today.replace(day=1)
    start_date = (first_day_of_month - datetime.timedelta(days=1)).replace(day=1) if today.day < 4 else first_day_of_month
    end_date = today - datetime.timedelta(days=1)
    return start_date, end_date

def get_db_connection(config):
    try:
        return pymssql.connect(config['server'], config['user'], config['password'], config['database'])
    except pymssql.DatabaseError as e:
        print(f"数据库连接失败: {e}")
        # 可以考虑返回 None 或抛出异常

def execute_query_and_process_data(conn, sql, decode_columns_list):
    try:
        with conn.cursor(as_dict=True) as cursor:
            cursor.execute(sql)
            data = cursor.fetchall()
            df = pd.DataFrame(data)
            # 如果需要解码
            if decode_columns_list:
                df = decode_columns(df, decode_columns_list)
            return df
    except pymssql.DatabaseError as e:
        print(f"查询执行失败: {e}")
        return None  

def write_df_to_excel(sheet, df, start_row, start_col, percent_columns=None):
    for index, row in enumerate(df.iterrows()):
        for j, value in enumerate(row[1]):
            cell = sheet.cell(row=start_row + index, column=start_col + j, value=value)
            if percent_columns and (start_col + j - 1) in percent_columns:
                cell.number_format = '0.0%'
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 定义计算比率的函数
def calculate_rates(df):
    df['消费会员企微添加率'] = df['总消费会员企微添加数'] / df['总消费人数']
    df['消费新客企微添加率'] = df['总新会员企微添加数'] / df['总新会员消费人数']
    df['消费老客企微添加率'] = df['总老会员企微添加数'] / df['总老会员消费人数']

    # 处理分母为0的情况
    df['消费会员企微添加率'].fillna(0, inplace=True)
    df['消费新客企微添加率'].fillna(0, inplace=True)
    df['消费老客企微添加率'].fillna(0, inplace=True)

    return df


# 计算每个大区的整体比率
def calculate_total_rates(group):
    total_consumers = group['总消费人数'].sum()
    total_new_members = group['总新会员消费人数'].sum()
    total_old_members = group['总老会员消费人数'].sum()

    total_added_members = group['总消费会员企微添加数'].sum()
    total_added_new_members = group['总新会员企微添加数'].sum()
    total_added_old_members = group['总老会员企微添加数'].sum()

    group['整体_消费会员企微添加率'] = total_added_members / total_consumers if total_consumers > 0 else 0
    group['整体_消费新客企微添加率'] = total_added_new_members / total_new_members if total_new_members > 0 else 0
    group['整体_消费老客企微添加率'] = total_added_old_members / total_old_members if total_old_members > 0 else 0
    return group

# 为每个门店计算整体比率
def calculate_total_store_rates(group):
    total_rates = {
        ('整体', '消费会员企微添加率'): group['消费会员企微添加率'].mean(),
        ('整体', '消费新客企微添加率'): group['消费新客企微添加率'].mean(),
        ('整体', '消费老客企微添加率'): group['消费老客企微添加率'].mean()
    }
    return pd.Series(total_rates)

start_date, end_date = get_dynamic_dates()
params = {"start_date":start_date,"end_date":end_date}
round_func = lambda x: round(x, 2)

# %% [markdown]
# ## 原始数据提取

# %%
db_config = {
    'server': '192.168.0.169',
    'user': 'chjreport',
    'password': 'Chj@12345',
    'database': 'ChjBidb'
}
sql = f'''
        SELECT 
            a.大区,
            a.小区,
            a.门店代码,
            a.门店名称,
            a.[门店类型],
            SUM(a.消费人数) AS 总消费人数,
            SUM(a.消费会员企微添加数) AS 总消费会员企微添加数,
            SUM(a.新会员消费人数) AS 总新会员消费人数,
            SUM(a.新会员企微添加数) AS 总新会员企微添加数,
            SUM(a.老会员消费人数) AS 总老会员消费人数,
            SUM(a.老会员企微添加数) AS 总老会员企微添加数
        FROM v_shop_dict s
        LEFT JOIN (
            SELECT 
                c.大区,
                c.小区,
                c.[门店类型],
                c.门店代码,
                c.门店名称,
                COUNT(DISTINCT c.会员号) AS 消费人数,
                COUNT(DISTINCT CASE WHEN m.注册时间 BETWEEN '{start_date}' AND '{end_date}'+' 23:59:59' THEN c.会员号 ELSE NULL END) AS 新会员消费人数,
                COUNT(DISTINCT CASE WHEN m.注册时间 < '{start_date}' THEN c.会员号 ELSE NULL END) AS 老会员消费人数,
                COUNT(DISTINCT q.unionId) AS 消费会员企微添加数,
                COUNT(DISTINCT CASE WHEN m.注册时间 BETWEEN '{start_date}' AND '{end_date}'+' 23:59:59' THEN q.unionId ELSE NULL END) AS 新会员企微添加数,
                COUNT(DISTINCT CASE WHEN m.注册时间 < '{start_date}' THEN q.unionId ELSE NULL END) AS 老会员企微添加数
            FROM bi_business_member AS m 
            JOIN bi_business_consume AS c ON m.会员号 = c.会员号 
            JOIN v_btgoods AS g ON g.商品代码 = c.商品代码 
            LEFT JOIN (SELECT DISTINCT 企微unionId AS unionId FROM BI_Business_qywechatmember_all WHERE 企微添加时间 <= '{end_date}'+' 23:59:59') q ON m.微信unionId = q.unionId
            WHERE c.来源 IN ('智慧云店','智能中台','中台') 
                AND c.消费时间 BETWEEN '{start_date}' AND '{end_date}'+' 23:59:59' 
                AND g.二级大类 NOT IN ('其他','促销物料') 
                AND c.金额 > 0 
                AND c.商品代码 NOT IN ('SHG30000905','Q3G30000105','QQG30000091','QQG30000090','SHG30000943','SHG30001192') 
                AND m.是否在云店上激活 = '是'
            GROUP BY c.大区, c.小区, c.[门店类型], c.门店代码, c.门店名称
        ) as a ON a.门店代码 = s.门店代码 
        LEFT JOIN (
            SELECT ltrim(rtrim(t.firstShopCode)) as 门店代码, count(distinct t.mobile) as 门店吸粉数
            FROM v_mall_member t
            WHERE t.date_add BETWEEN '{start_date}' AND '{end_date}'+' 23:59:59'
                AND t.firstguideno NOT IN ('84e67d1115c859c336','2f7e946583048a96fe')  
            GROUP BY ltrim(rtrim(t.firstShopCode))
        ) as b ON s.门店代码 = b.门店代码
        WHERE (a.门店代码 IS NOT NULL OR b.门店代码 IS NOT NULL) and 大区 is not null and 大区 != '营销管理部'
        GROUP BY a.大区, a.小区, a.门店名称, a.[门店类型], a.门店代码
'''
task_details = [
    {"task_name": "消费7天", "decode_columns_list": ["大区","小区","门店名称","门店类型","门店代码"]}
]
with get_db_connection(db_config) as conn:
    for task in task_details:
        result = execute_query_and_process_data(conn, sql, task["decode_columns_list"])

# %% [markdown]
# ## 门店明细

# %%
# 对原始数据按照门店类型分组，并计算每组的比率
grouped_result = result.groupby('门店类型',group_keys=False).apply(calculate_rates)
grouped_result
order = ['大区','小区','门店代码','门店名称','门店类型', '总消费人数','总消费会员企微添加数', '消费会员企微添加率', '总新会员消费人数', '总新会员企微添加数', '消费新客企微添加率', '总老会员消费人数','总老会员企微添加数', '消费老客企微添加率']
# # 请根据您实际的列名进行调整
sorted_result = grouped_result[order]

# %% [markdown]
# ## 大区添加率

# %%
# 直接计算整个DataFrame的比率
result_daqu = calculate_rates(result)
# 按大区和门店类型分组，计算平均值
grouped = result_daqu.groupby(['大区', '门店类型']).agg({'消费会员企微添加率': 'mean', '消费新客企微添加率': 'mean', '消费老客企微添加率': 'mean'}).unstack(fill_value=0)


# 应用整体比率计算
total_rates = result_daqu.groupby('大区',group_keys=False).apply(calculate_total_rates).groupby('大区').agg({'整体_消费会员企微添加率': 'mean', '整体_消费新客企微添加率': 'mean', '整体_消费老客企微添加率': 'mean'})

# 合并分组结果和整体比率
grouped = pd.concat([grouped, total_rates], axis=1)


# 首先，将"整体"指标列转换为多级索引格式
grouped[('整体', '消费会员企微添加率')] = grouped['整体_消费会员企微添加率']
grouped[('整体', '消费新客企微添加率')] = grouped['整体_消费新客企微添加率']
grouped[('整体', '消费老客企微添加率')] = grouped['整体_消费老客企微添加率']

# 然后，使用reindex重新排列列
new_columns = [
    ('整体', '消费会员企微添加率'),
    ('整体', '消费新客企微添加率'),
    ('整体', '消费老客企微添加率'),
    ('消费会员企微添加率', '自营'),
    ('消费新客企微添加率', '自营'),
    ('消费老客企微添加率', '自营'),
    ('消费会员企微添加率', '代理'),
    ('消费新客企微添加率', '代理'),
    ('消费老客企微添加率', '代理')
]

grouped = grouped.reindex(columns=new_columns)
grouped.reset_index(inplace=True)
grouped = grouped.sort_values(by=('整体', '消费会员企微添加率'), ascending=True)
# 对自营门店数据进行筛选
result_self_operated = result[result['门店类型'] == '自营']

# 对代理门店数据进行筛选
result_agent = result[result['门店类型'] == '代理']

# 计算自营门店的比率
total_data_self_operated = result_self_operated.agg({
    '总消费人数': 'sum',
    '总消费会员企微添加数': 'sum',
    '总新会员消费人数': 'sum',
    '总新会员企微添加数': 'sum',
    '总老会员消费人数': 'sum',
    '总老会员企微添加数': 'sum'
})

# 计算代理门店的比率
total_data_agent = result_agent.agg({
    '总消费人数': 'sum',
    '总消费会员企微添加数': 'sum',
    '总新会员消费人数': 'sum',
    '总新会员企微添加数': 'sum',
    '总老会员消费人数': 'sum',
    '总老会员企微添加数': 'sum'
})
total_data = result.agg({
    '总消费人数': 'sum',
    '总消费会员企微添加数': 'sum',
    '总新会员消费人数': 'sum',
    '总新会员企微添加数': 'sum',
    '总老会员消费人数': 'sum',
    '总老会员企微添加数': 'sum'
})
# 计算全国水平的比率
total_rates = {
    ('整体', '消费会员企微添加率'): total_data['总消费会员企微添加数'] / total_data['总消费人数'] if total_data['总消费人数'] > 0 else 0,
    ('整体', '消费新客企微添加率'): total_data['总新会员企微添加数'] / total_data['总新会员消费人数'] if total_data['总新会员消费人数'] > 0 else 0,
    ('整体', '消费老客企微添加率'): total_data['总老会员企微添加数'] / total_data['总老会员消费人数'] if total_data['总老会员消费人数'] > 0 else 0,
    ('消费会员企微添加率', '自营'): total_data_self_operated['总消费会员企微添加数'] / total_data_self_operated['总消费人数'] if total_data_self_operated['总消费人数'] > 0 else 0,
    ('消费新客企微添加率', '自营'): total_data_self_operated['总新会员企微添加数'] / total_data_self_operated['总新会员消费人数'] if total_data_self_operated['总新会员消费人数'] > 0 else 0,
    ('消费老客企微添加率', '自营'): total_data_self_operated['总老会员企微添加数'] / total_data_self_operated['总老会员消费人数'] if total_data_self_operated['总老会员消费人数'] > 0 else 0,
    ('消费会员企微添加率', '代理'): total_data_agent['总消费会员企微添加数'] / total_data_agent['总消费人数'] if total_data_agent['总消费人数'] > 0 else 0,
    ('消费新客企微添加率', '代理'): total_data_agent['总新会员企微添加数'] / total_data_agent['总新会员消费人数'] if total_data_agent['总新会员消费人数'] > 0 else 0,
    ('消费老客企微添加率', '代理'): total_data_agent['总老会员企微添加数'] / total_data_agent['总老会员消费人数'] if total_data_agent['总老会员消费人数'] > 0 else 0
}

# 创建表示全国水平的行
national_row = pd.DataFrame([total_rates], index=["全国"])

# 将全国水平的行添加到grouped DataFrame中
grouped = pd.concat([grouped, national_row])


# %% [markdown]
# ## 小区添加率

# %%
# 应用计算比率的函数
result_xiaoqu = calculate_rates(result)
# 按大区和门店类型分组，计算平均值
grouped2 = result_xiaoqu.groupby(['小区', '门店类型'], group_keys=False).agg({'消费会员企微添加率': 'mean', '消费新客企微添加率': 'mean', '消费老客企微添加率': 'mean'}).unstack(fill_value=0)


# 应用整体比率计算
total_rates = result_xiaoqu.groupby('小区', group_keys=False).apply(calculate_total_rates).groupby('小区').agg({'整体_消费会员企微添加率': 'mean', '整体_消费新客企微添加率': 'mean', '整体_消费老客企微添加率': 'mean'})

# 合并分组结果和整体比率
grouped2 = pd.concat([grouped2, total_rates], axis=1)
# 首先，将"整体"指标列转换为多级索引格式
grouped2[('整体', '消费会员企微添加率')] = grouped2['整体_消费会员企微添加率']
grouped2[('整体', '消费新客企微添加率')] = grouped2['整体_消费新客企微添加率']
grouped2[('整体', '消费老客企微添加率')] = grouped2['整体_消费老客企微添加率']
# 然后，使用reindex重新排列列
new_columns = [
    ('整体', '消费会员企微添加率'),
    ('整体', '消费新客企微添加率'),
    ('整体', '消费老客企微添加率'),
    ('消费会员企微添加率', '自营'),
    ('消费新客企微添加率', '自营'),
    ('消费老客企微添加率', '自营'),
    ('消费会员企微添加率', '代理'),
    ('消费新客企微添加率', '代理'),
    ('消费老客企微添加率', '代理')
]
grouped2 = grouped2.reindex(columns=new_columns)
# 去除索引列
grouped2.reset_index(inplace=True)
grouped2
# 按照整体_消费会员企微添加率从小到大排序，取20个
grouped2 = grouped2.sort_values(by=('整体', '消费会员企微添加率'), ascending=True).head(20)

# %%
result2  =calculate_rates(result)
# 按门店名称和门店类型分组，计算平均值
grouped4 = result2.groupby(['门店名称', '门店类型']).agg({'消费会员企微添加率': 'mean', '消费新客企微添加率': 'mean', '消费老客企微添加率': 'mean'}).unstack(fill_value=0)
total_rates = result2.groupby('门店名称').apply(calculate_total_store_rates)
# 将整体比率添加到grouped4中，并保持列的多级索引结构
grouped4 = pd.concat([grouped4, total_rates], axis=1)
# 重新排列列，确保使用正确的多级索引格式
columns_order = [
    ('整体', '消费会员企微添加率'),
    ('整体', '消费新客企微添加率'),
    ('整体', '消费老客企微添加率'),
    ('消费会员企微添加率', '自营'),
    ('消费新客企微添加率', '自营'),
    ('消费老客企微添加率', '自营'),
    ('消费会员企微添加率', '代理'),
    ('消费新客企微添加率', '代理'),
    ('消费老客企微添加率', '代理')
]
grouped4 = grouped4[columns_order]
# 重置索引并展示
grouped4.reset_index(inplace=True)

# %% [markdown]
# ## 导出文档

# %%
# 加载现有的 Excel 文件
file_path = r'./企微添加率监控测试1.0.xlsx'
book = load_workbook(file_path)
# 大区统计表
sheet1 = book['大区情况'] 
write_df_to_excel(sheet1, grouped, start_row=4, start_col=1, percent_columns=range(2, sheet1.max_column + 1))
# 落后小区
sheet2 = book['落后小区']
write_df_to_excel(sheet2, grouped2, start_row=4, start_col=1, percent_columns=range(2, sheet2.max_column + 1))

# 门店情况
sheet3 = book['门店情况']
write_df_to_excel(sheet3, grouped4, start_row=4, start_col=1, percent_columns=range(12, sheet3.max_column + 1))

# 数据明细
sheet4 = book['数据明细']
write_df_to_excel(sheet4, sorted_result, start_row=2, start_col=1, percent_columns=[7,10,13])

# 保存更改
book.save(file_path)
print('写入成功')


